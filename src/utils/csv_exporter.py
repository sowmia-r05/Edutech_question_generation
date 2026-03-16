"""Excel exporter — writes questions into the EXACT Quiz Upload Template format.

Template: Quiz_Upload_Template__1_.xlsx
Sheet:    Questions
Row 1:    Headers  (fill: #4338CA, font: white bold 11pt, center aligned)
Row 2:    Instructions (fill: #EEF2FF, font: Calibri)
Row 3+:   Data rows   (fill: #EFF6FF, font: Calibri, wrap text)

Columns:
  A: question_text  ← plain question text only (no image HTML)
  B: type
  C: option_a
  D: option_b
  E: option_c
  F: option_d
  G: correct_answer
  H: points
  I: category
  J: image_url      ← base64 data URI compressed to <20KB (falls back to S3 URL)
  K: explanation
"""

import base64
import io
import logging
import os
from datetime import datetime
from pathlib import Path
from shutil import copyfile

logger = logging.getLogger(__name__)

# Max image size in bytes before compression (20 KB)
MAX_IMAGE_BYTES = 20 * 1024


def compress_image_to_limit(image_bytes: bytes, max_bytes: int = MAX_IMAGE_BYTES) -> bytes:
    """
    Compress a PNG/JPEG image to stay under max_bytes using PIL.
    Tries progressively lower quality until the target is met.
    Returns compressed bytes (PNG if original was PNG, JPEG otherwise).
    """
    try:
        from PIL import Image
    except ImportError:
        logger.warning("Pillow not installed — image not compressed. Run: pip install Pillow")
        return image_bytes

    if len(image_bytes) <= max_bytes:
        return image_bytes

    img = Image.open(io.BytesIO(image_bytes)).convert("RGB")

    # Step 1: Shrink dimensions progressively
    for scale in [0.8, 0.6, 0.5, 0.4, 0.3, 0.25, 0.2]:
        w = int(img.width * scale)
        h = int(img.height * scale)
        resized = img.resize((max(w, 60), max(h, 60)), Image.LANCZOS)

        # Step 2: Try JPEG at decreasing quality
        for quality in [85, 70, 55, 40, 30, 20]:
            buf = io.BytesIO()
            resized.save(buf, format="JPEG", quality=quality, optimize=True)
            data = buf.getvalue()
            if len(data) <= max_bytes:
                logger.info(
                    f"Image compressed: {len(image_bytes)//1024}KB → "
                    f"{len(data)//1024}KB  (scale={scale}, quality={quality})"
                )
                return data

    # Last resort — very small JPEG
    buf = io.BytesIO()
    img.resize((80, 60)).save(buf, format="JPEG", quality=15, optimize=True)
    data = buf.getvalue()
    logger.warning(f"Image aggressively compressed to {len(data)} bytes")
    return data


class CSVExporter:
    """Export questions into the exact Quiz Upload Template xlsx format."""

    HEADER_ROW     = 1
    INST_ROW       = 2
    DATA_START_ROW = 3

    # Exact column headers from the template
    HEADERS = [
        "question_text",
        "type",
        "option_a",
        "option_b",
        "option_c",
        "option_d",
        "correct_answer",
        "points",
        "category",
        "image_url",
        "explanation",
    ]

    INSTRUCTIONS = [
        "The question text (REQUIRED)",
        "radio_button | checkbox | free_text | picture_choice",
        "Option A text",
        "Option B text",
        "Option C text (optional)",
        "Option D text (optional)",
        "Letter: A, B, C, D (comma-separate for checkbox)",
        "Points (default: 1)",
        "Topic — e.g. Fractions, Grammar",
        "URL to question image (optional)",
        "Explanation shown after quiz (optional)",
    ]

    # Exact colors from template inspection
    _HDR_FILL  = "FF4338CA"   # Row 1 — deep indigo
    _INST_FILL = "FFEEF2FF"   # Row 2 — light lavender
    _DATA_FILL = "FFEFF6FF"   # Row 3+ — very light blue

    # Column widths matching the template exactly
    _COL_WIDTHS = {
        "A": 50.0, "B": 18.0, "C": 22.0, "D": 22.0, "E": 22.0,
        "F": 22.0, "G": 16.0, "H": 10.0, "I": 22.0, "J": 30.0, "K": 40.0,
    }

    # Row heights matching the template
    _HDR_HEIGHT  = 36.0
    _INST_HEIGHT = 30.0
    _DATA_HEIGHT = 32.0

    # ── Sub-subject → Major Topic lookup sets ─────────────────────────────────

    _NUMBER_ALGEBRA_SUBS = {
        "counting objects", "reading numbers", "ordering numbers",
        "tens and ones", "hundreds", "value of digits",
        "simple addition", "addition with carrying",
        "simple subtraction", "subtraction with regrouping",
        "repeated addition", "groups of objects", "simple multiplication facts",
        "sharing equally", "grouping objects",
        "increasing patterns", "skip counting", "missing numbers",
        "odd and even numbers", "half", "quarter", "equal parts",
        "adding money", "subtracting money", "counting coins",
        "place value", "addition", "subtraction", "multiplication",
        "division", "number patterns", "fractions", "money",
        "counting and number recognition", "multiplication groups and arrays",
        "division sharing", "simple fractions", "number and place value",
        "patterns and algebra", "money and financial mathematics",
        "number", "algebra", "arithmetic",
    }

    _MEASUREMENT_GEOMETRY_SUBS = {
        "length", "mass", "capacity", "time", "area",
        "2d shapes", "3d objects", "position and direction",
        "measurement", "geometry", "shapes", "angles", "symmetry",
        "perimeter", "volume", "2d shape", "3d object",
        "position", "direction",
    }

    _STATISTICS_PROBABILITY_SUBS = {
        "reading graphs", "picture graphs", "bar graphs",
        "data interpretation", "chance likely unlikely",
        "data", "statistics", "graphs", "tables",
        "probability", "chance", "data and statistics",
        "tally", "pictograph", "survey", "column graph",
    }

    _DIFFICULTY_LABELS = {
        0: "Easy", 1: "Easy", 2: "Medium", 3: "Medium", 4: "Hard", 5: "Hard"
    }

    _CORRECT_LETTER_MAP = {0: "A", 1: "B", 2: "C", 3: "D"}

    # ─────────────────────────────────────────────────────────────────────────

    def export_to_xlsx(
        self,
        questions,
        output_path: str | None = None,
        grade: str = "grade",
        template_path: str | None = None,
    ) -> str:
        """
        Write questions into the Quiz Upload Template xlsx format.
        Uses the exact template file to preserve all styling.
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            raise ImportError("Run: pip install openpyxl")

        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            os.makedirs("output", exist_ok=True)
            output_path = f"output/questions_{grade}_{timestamp}.xlsx"

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        # ── Find and copy the EXACT template file ─────────────────────────────
        template_found = None
        candidates = []
        if template_path:
            candidates.append(template_path)
        candidates += [
            "Quiz_Upload_Template__1_.xlsx",
            "Quiz_Upload_Template.xlsx",
        ]
        for c in candidates:
            if Path(c).exists():
                template_found = c
                break

        if template_found:
            # Copy template → output path to preserve ALL styles exactly
            copyfile(template_found, output_path)
            wb = openpyxl.load_workbook(output_path)
            ws = wb["Questions"]

            # Clear only the example data rows (row 3 onwards)
            for row in ws.iter_rows(min_row=self.DATA_START_ROW, max_row=ws.max_row):
                for cell in row:
                    cell.value = None

            logger.info(f"Using exact template: {template_found}")
        else:
            # Build from scratch matching the template exactly
            logger.warning(
                "Quiz_Upload_Template__1_.xlsx not found in project root. "
                "Copy it there for pixel-perfect output. Building from scratch."
            )
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Questions"

            hdr_fill  = PatternFill("solid", fgColor=self._HDR_FILL)
            hdr_font  = Font(name="Calibri", bold=True, color="FFFFFFFF", size=11)
            hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

            inst_fill  = PatternFill("solid", fgColor=self._INST_FILL)
            inst_font  = Font(name="Calibri", size=10)
            inst_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

            for col_idx, (h, inst) in enumerate(
                zip(self.HEADERS, self.INSTRUCTIONS), start=1
            ):
                c1 = ws.cell(row=1, column=col_idx, value=h)
                c1.font = hdr_font; c1.fill = hdr_fill; c1.alignment = hdr_align
                c2 = ws.cell(row=2, column=col_idx, value=inst)
                c2.font = inst_font; c2.fill = inst_fill; c2.alignment = inst_align

            for col_letter, width in self._COL_WIDTHS.items():
                ws.column_dimensions[col_letter].width = width

            ws.row_dimensions[1].height = self._HDR_HEIGHT
            ws.row_dimensions[2].height = self._INST_HEIGHT

        # ── Write question data starting at row 3 ────────────────────────────
        try:
            from openpyxl.styles import Alignment, Font, PatternFill
            data_fill  = PatternFill("solid", fgColor=self._DATA_FILL)
            data_font  = Font(name="Calibri", size=11)
            data_align = Alignment(vertical="center", wrap_text=True)
        except Exception:
            data_fill  = None
            data_font  = None
            data_align = None

        for q_idx, q in enumerate(questions):
            row_num  = self.DATA_START_ROW + q_idx
            row_data = self._build_row(q)

            ws.row_dimensions[row_num].height = self._DATA_HEIGHT

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                if data_fill:
                    cell.fill      = data_fill
                    cell.font      = data_font
                    cell.alignment = data_align

        wb.save(output_path)
        logger.info(f"Exported {len(questions)} questions to: {output_path}")
        return output_path

    def export_to_csv(self, questions, output_path=None, grade="grade") -> str:
        import csv
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            os.makedirs("output", exist_ok=True)
            output_path = f"output/questions_{grade}_{timestamp}.csv"

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(self.HEADERS)
            for q in questions:
                writer.writerow(self._build_row(q))

        logger.info(f"Exported {len(questions)} questions to CSV: {output_path}")
        return output_path

    @classmethod
    def export_with_summary(cls, questions, output_dir, base_filename) -> dict:
        exporter = cls()
        output_path = str(Path(output_dir) / f"{base_filename}.xlsx")
        out = exporter.export_to_xlsx(questions=questions, output_path=output_path)
        return {"csv": out, "summary": out}

    # ── Row builder ───────────────────────────────────────────────────────────

    def _build_row(self, q) -> list:
        """
        Build one data row.

        question_text (column A): always plain text — no image HTML embedded.
        image_url (column J):     base64 data URI (data:image/jpeg;base64,...),
                                  compressed to <20KB. Falls back to S3 URL if
                                  base64 not available.
        """
        options = list(getattr(q, "options", []) or [])
        while len(options) < 4:
            options.append("")
        options = [self._clean_option(opt) for opt in options[:4]]

        correct_idx    = self._get_correct_index(q)
        correct_letter = self._CORRECT_LETTER_MAP.get(correct_idx, "A")

        explanation = getattr(q, "explanation", "") or ""
        category    = self._build_categories(q)

        # ── image_url (column J): use compressed base64 data URI ─────────────
        # q.image_base64 is set by generate_questions.py after downloading the
        # S3 image, compressing it to <20KB, and encoding as base64.
        image_url = getattr(q, "image_base64", None) or ""
        if image_url and not image_url.startswith("data:"):
            image_url = f"data:image/jpeg;base64,{image_url}"

        # Fall back to S3 URL if no base64 available
        if not image_url:
            artifacts = getattr(q, "artifacts", None)
            if artifacts and isinstance(artifacts, list) and artifacts:
                image_url = str(artifacts[0])
        if not image_url:
            image_url = getattr(q, "question_image", "") or ""

        # ── question_text (column A): always plain text ───────────────────────
        question_text = q.question_text

        return [
            question_text,      # A — question_text (plain, no image HTML)
            "radio_button",     # B — type
            options[0],         # C — option_a
            options[1],         # D — option_b
            options[2],         # E — option_c
            options[3],         # F — option_d
            correct_letter,     # G — correct_answer
            1,                  # H — points
            category,           # I — category
            image_url,          # J — image_url (base64 data URI or S3 URL)
            explanation,        # K — explanation
        ]

    # ── Category builder ──────────────────────────────────────────────────────

    def _build_categories(self, q) -> str:
        """Returns: "Major Topic, Sub Topic, Difficulty"  e.g. "Number & Algebra, Place Value, Easy" """
        sub       = (getattr(q, "sub_subject", "") or "").strip()
        sub_lower = sub.lower()

        if sub_lower in self._NUMBER_ALGEBRA_SUBS:
            major = "Number & Algebra"
        elif sub_lower in self._MEASUREMENT_GEOMETRY_SUBS:
            major = "Measurement & Geometry"
        elif sub_lower in self._STATISTICS_PROBABILITY_SUBS:
            major = "Statistics & Probability"
        else:
            if any(k in sub_lower for k in [
                "count", "number", "add", "subtract", "multipl", "divis",
                "fraction", "money", "pattern", "place", "odd", "even",
                "coin", "skip", "algebra", "arith",
            ]):
                major = "Number & Algebra"
            elif any(k in sub_lower for k in [
                "length", "mass", "time", "area", "shape", "object",
                "position", "capacity", "geometry", "measure", "perimeter",
                "volume", "angle", "symmetr",
            ]):
                major = "Measurement & Geometry"
            elif any(k in sub_lower for k in [
                "graph", "data", "chance", "probab", "statistic",
                "tally", "survey", "pictograph",
            ]):
                major = "Statistics & Probability"
            else:
                major = "Numeracy"

        sub_display = sub.title() if sub else "General"

        diff_int = getattr(q, "difficulty", 0)
        try:
            diff_int = int(diff_int)
        except (TypeError, ValueError):
            diff_int = 0
        diff_label = self._DIFFICULTY_LABELS.get(diff_int, "Easy")

        return f"{major}, {sub_display}, {diff_label}"

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _get_correct_index(self, q) -> int:
        val = getattr(q, "correct_option_index", None)
        if val is not None:
            return int(val)
        val = getattr(q, "answer_index", None)
        if val is not None:
            return int(val)
        val = getattr(q, "correct_answer", None)
        if val is not None:
            return {"A": 0, "B": 1, "C": 2, "D": 3}.get(str(val).upper(), 0)
        return 0

    @staticmethod
    def _clean_option(option_text: str) -> str:
        if not option_text:
            return ""
        text = str(option_text).strip()
        if len(text) > 2 and text[1] in (".", ")") and text[0].upper() in "ABCD1234":
            return text[2:].strip()
        return text